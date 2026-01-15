import os
import random

import gradio as gr
import openpyxl
import openpyxl.cell
import ujson as json
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

from utils import read_json, replace_wildcards, sleep_for_cool
from utils.environment import env
from utils.generator import Generator
from utils.logger import logger

generator = Generator("https://image.novelai.net/ai/generate-image")


def number_to_letters(n):
    result = ""
    while n >= 0:
        result = chr(n % 26 + ord("A")) + result
        n = n // 26 - 1
    return result


if not os.path.exists(
    xlsx_path := "./plugins/anr_plugin_stories_to_images/模板文件.xlsx"
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.row_dimensions[1].height = 50
    for row in range(2, 999):
        ws.row_dimensions[row].height = 300
    for col in [number_to_letters(num) for num in range(1000)]:
        ws.column_dimensions[col].width = 40
    ws.append(
        [
            "推文",
            "TAG",
            "图片",
        ]
    )
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
    wb.save(xlsx_path)


def open_folder(path):
    os.startfile(os.path.abspath(path))


def main(file_path, images_number):
    with open("./outputs/temp_break.json", "w") as f:
        json.dump({"break": False}, f)

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Sheet"]
    col_num = 2
    positive = sheet[f"B{col_num}"].value

    try:
        json_data = read_json("./last.json")
    except FileNotFoundError:
        logger.error("未进行一次图片生成!")
        return gr.update(value="未进行一次图片生成!", visible=True)

    num = 1
    while positive is not None:
        _break = read_json("./outputs/temp_break.json")
        if _break["break"]:
            logger.warning("已停止生成!")
            break

        row_num = ord("C") - 65
        for _ in range(images_number):
            _break = read_json("./outputs/temp_break.json")
            if _break["break"]:
                break

            logger.info(f"正在生成第 {num} 张图片...")
            _positive = replace_wildcards(positive)
            if json_data.get("model") in ["nai-diffusion-3", "nai-diffusion-furry-3"]:
                pass
            else:
                json_data["parameters"]["v4_prompt"]["caption"][
                    "base_caption"
                ] = _positive
            json_data["input"] = _positive
            json_data["parameters"]["seed"] = random.randint(1000000000, 9999999999)

            saved_path = None
            while saved_path is None:
                _break = read_json("./outputs/temp_break.json")
                if _break["break"]:
                    logger.warning("已停止生成!")
                    break

                image_data = generator.generate(json_data)
                if image_data:
                    saved_path = generator.save(
                        image_data, "text2image", json_data["parameters"]["seed"]
                    )
                else:
                    sleep_for_cool(5)
            num += 1

            image = Image(saved_path)
            w = image.width
            h = image.height
            image.width, image.height = 265, int(265 / w * h)
            sheet.add_image(image, "{}{}".format(number_to_letters(row_num), col_num))

            sleep_for_cool(env.cool_time)

            row_num += 1
        col_num += 1
        positive = sheet[f"B{col_num}"].value

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    workbook.save(file_path)
    logger.success(f"全部 TAG 生成完毕, 打开 {file_path} 以查看结果!")

    return gr.update(
        value=f"全部 TAG 生成完毕, 打开 {file_path} 以查看结果!", visible=True
    )
